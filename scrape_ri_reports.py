import concurrent.futures
import os
import re
import zipfile
import openpyxl
import pandas as pd
import numpy as np
from openpyxl.utils import get_column_interval
import random
from openpyxl.utils.exceptions import InvalidFileException
import sqlalchemy as sa
from itertools import repeat
from sqlalchemy import create_engine, inspect, text
from sqlalchemy import types
from sqlalchemy import exc



pd.set_option('display.max_columns', None)

special_path_2023 = r'\\ad.trw.com\reynosa2\Departments\QM\Receiving Inspection\05_Receiving Inspection\18-Reportes de RI\Files-anteriores\Matriz Especiales 2023\Reportes Excel'
special_path_2022 = r'\\ad.trw.com\reynosa2\Departments\QM\Receiving Inspection\05_Receiving Inspection\18-Reportes ' \
                    r'de RI\Files-anteriores\Matriz Especiales\Reportes Excel'

# Todo. Add a way to edit the month.
normal_tests_path = r'\\ad.trw.com\reynosa2\Departments\QM\Receiving Inspection\05_Receiving ' \
                    r'Inspection\18-Reportes de RI\Resultados de Pruebas de Proceo\2023\MARZO'
special_register_path = r'\\ad.trw.com\reynosa2\Departments\QM\Receiving Inspection\05_Receiving ' \
                        r'Inspection\18-Reportes de RI\Files-anteriores\Matriz Especiales 2023\\Registro pruebas ' \
                        r'de metrologia 2023.xlsx'
output_path = r'C:\Repos\CSPM\RI\Output Files'

log_dtypes = {
    'TRWNumber': types.VARCHAR(length=50),
    'review': types.VARCHAR(length=50),
    'lote': types.VARCHAR(length=50),
    'ri_log_id': types.BIGINT(),
    'supplier': types.VARCHAR(length=50),
    'quantity': types.SMALLINT(),
    'log_date': types.DATE(),
    'technician': types.VARCHAR(length=50),
    'line': types.VARCHAR(length=50),
    'shift': types.SMALLINT(),
    'log_test': types.VARCHAR(length=50),
    'comments': types.VARCHAR(),
    'kind': types.VARCHAR(length=50),
}

tests_dtypes = {
    'sample_batch': types.SMALLINT(),
    'sample_number': types.SMALLINT(),
    'qa': types.VARCHAR(length=30),
    'characteristic': types.VARCHAR(length=30),
    'nominal_value': types.FLOAT(),
    'superior_limit': types.FLOAT(),
    'inferior_limit': types.FLOAT(),
    'equipment': types.VARCHAR(length=30),
    'sample_value': types.FLOAT(),
    'ri_log_id': types.BIGINT(),
}


# * Specify the connection details
SERVER = 'zam-SPC-rey-db'
DATABASE = 'SPC-Rey'
conn_string = f'mssql+pyodbc://@{SERVER}/{DATABASE}?trusted_connection=yes&driver=ODBC+Driver+17+for+SQL+Server'
sql_engine = create_engine(conn_string, fast_executemany=True)
sql_inspector = inspect(sql_engine)


def xlrange_to_df_drop_empty(range_string, ws):
    """
    Load a range of data from a worksheet into a Pandas DataFrame but does not keep the empty cells.

    Parameters:
    range_string (str): A string representing the range of cells to be loaded, in Excel notation (e.g. "A1:B10").
    ws (openpyxl.Worksheet): The worksheet to load the data from.

    Returns:
    pandas.DataFrame: A DataFrame containing the loaded data.

    Raises:
    ValueError: If the specified range is not valid.
    """

    col_start, col_end = re.findall(r"[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    df = pd.DataFrame(data_rows)
    df.columns = get_column_interval(col_start, col_end)

    df.dropna(axis=1, inplace=True)

    return df


def conc_samples_to_df(xl_range, phase, ws):
    """
    Converts a range of concentricity samples from an Excel worksheet to a Pandas DataFrame.

    Parameters:
    xl_range (str): A string representing the range of cells to be loaded, in Excel notation (e.g. "C5:M15").
    phase (str): A string representing the phase of the samples (e.g. "Phase 1").
    lot (str): A string representing the tube lot of the samples (e.g. "A1").
    ws (openpyxl.Worksheet): The worksheet to load the data from.

    Returns:
    pandas.DataFrame: A DataFrame containing the loaded samples, with additional columns for Tube Lot and Fase.

    Raises:
    ValueError: If the specified range is not valid.
    """
    # Gets the start samples.
    conc_samples = xlrange_to_df_drop_empty(xl_range, ws)
    conc_samples['sample_batch'] = phase

    if len(conc_samples.columns) < 11:
        return None  # ? Before pd.DataFrame(columns=['sample_number', 'sample_value', 'sample_batch'])

    conc_samples.rename(columns={"C": "sample_number", 'M': "sample_value"}, inplace=True)

    # Todo. Find a way to append this mf into another fucking table
    # ! shit.
    conc_samples.drop(['D', 'E', 'F', 'G', 'I', 'J', 'K', 'L'], axis=1, inplace=True)

    return conc_samples


def xlrange_to_df_keep_empty(xlrange, worksheet):
    """Converts an Excel range to a pandas DataFrame while preserving empty cells.

    Args:
        worksheet (str): The worksheet containing the range.
        xlrange (str): The Excel range to convert (e.g. "A1:D10").

    Returns:
        pandas.DataFrame: The pandas DataFrame representing the Excel range.
    """

    # Get the values and dimensions of the range
    # noinspection PyTypeChecker
    rows = worksheet[xlrange]
    values = [[cell.value for cell in row] for row in rows]
    num_rows, num_cols = len(values), len(values[0])

    # Create the pandas DataFrame
    df = pd.DataFrame(index=range(num_rows), columns=range(num_cols))
    for row in range(num_rows):
        for col in range(num_cols):
            df.iat[row, col] = values[row][col]

    return df


def xlrange_to_list(range_string: str, ws):
    # Initialize an empty list
    values = []

    # Iterate over the rows and columns of the range
    for row in ws[range_string]:
        values = []
        for cell in row:
            values.append(cell.value)
        values.append(values[0])

    # Return the resulting list
    return values


def get_log_info(ws, prueba, kind):
    log_columns = ["Num. de Parte:", 'Revision:', 'Lote #:', 'Log #:', "Proveedor:", "Cantidad:", "Fecha:", "Tecnico:",
                   "Linea:", "Turno:"]

    # ! special: R2 and R3. for normal S2 and S3 in the last two elements.
    if kind == 'normal':
        log_xlpos = ['C2', 'C3', 'G2', 'G3', 'K2', 'K3', 'N2', 'N3', 'S2', 'S3']
    else:
        log_xlpos = ['C2', 'C3', 'G2', 'G3', 'K2', 'K3', 'N2', 'N3', 'R2', 'R3']

    log_values = [ws[xlpos].value for xlpos in log_xlpos]
    log_df = pd.DataFrame(columns=log_columns)
    log_df.loc[len(log_df)] = log_values
    log_df['Prueba'] = prueba
    log_df['comments'] = ws['C47'].value

    log_df.rename(columns={
        "Num. de Parte:": "TRWNumber",
        "Revision:": "review",
        "Lote #:": "lote",
        "Log #:": "ri_log_id",
        "Proveedor:": "supplier",
        "Cantidad:": "quantity",
        "Fecha:": "log_date",
        "Tecnico:": "technician",
        "Linea:": "line",
        "Turno:": "shift",
        "Prueba": "log_test"
    }, inplace=True)

    return log_df


def read_concentricity(ws, kind):
    # * Gets the log values as a pd.Dataframe called log_df
    log_df = get_log_info(ws, 'Concentricity', kind)

    # * Gets the samples.
    start_samples = conc_samples_to_df("C14:M19", 0, ws)
    inter_samples = conc_samples_to_df("C21:M26", 1, ws)
    final_samples = conc_samples_to_df("C28:M33", 2, ws)

    if all(value is None for value in [start_samples, inter_samples, final_samples]):
        return None, None

    conc_samples = pd.concat([start_samples, inter_samples, final_samples], ignore_index=True)
    conc_samples['ri_log_id'] = log_df['ri_log_id'].iloc[0]

    conc_samples['superior_limit'] = 0.28
    conc_samples['inferior_limit'] = 0
    conc_samples['nominal_value'] = 0
    conc_samples['equipment'] = "ROM"
    conc_samples['characteristic'] = 'CONC'
    conc_samples['qa'] = "CONC"

    return log_df, conc_samples


def read_ovality(ws, kind):
    oval_samples = pd.DataFrame(columns=["sample_number", "sample_value", "sample_batch", "ri_log_id"])
    sample_xlranges = ["C14:D19", "C21:D26", "C28:D33", "O14:P19", "O21:P26", "O28:P33"]
    sample_phases = [0, 1, 2, 0, 1, 2]

    # * Gets the log info.
    log_df = get_log_info(ws, 'Ovality', kind)

    # * Gets the samples.
    for i, xl_range in enumerate(sample_xlranges):
        df = xlrange_to_df_drop_empty(xl_range, ws)
        if len(df.columns) > 1:
            df.rename(columns={df.columns[0]: 'sample_number', df.columns[1]: 'sample_value'}, inplace=True)
            df["sample_batch"] = sample_phases[i]
            df["ri_log_id"] = log_df['ri_log_id'].iloc[0]

            oval_samples = pd.concat([oval_samples, df], ignore_index=True)

    oval_samples['superior_limit'] = 0.3
    oval_samples['inferior_limit'] = 0
    oval_samples['nominal_value'] = 0
    oval_samples['equipment'] = "ROM"
    oval_samples['characteristic'] = 'OVAL'
    oval_samples['qa'] = "OVAL"

    return log_df, oval_samples


def read_normal(ws, kind):
    log_df = get_log_info(ws, 'Normal', kind)
    df = xlrange_to_df_keep_empty('C5:Z40', ws)

    df = df[df[0].notna()]
    df.reset_index(inplace=True, drop=True)

    # * It's a list containing all the first appearance of the QA#
    qa_indexes = df[df[0].fillna('').str.contains('QA', na=False)].index.tolist()

    # If there's two rows of qas...
    if len(qa_indexes) == 2:
        # Slice the DataFrame using loc up to the QA#, but not including it.
        row = df.loc[:qa_indexes[1] - 1, :]
        row2 = df.loc[qa_indexes[1]:, :]

        # Separate samples from qas
        qas = row.iloc[:6, 1::2]
        normal_samples = row.iloc[6:, :].dropna(axis=1, how='all')
        qas2 = row2.iloc[:6, 1::2]
        samples2 = row2.iloc[6:, :].dropna(axis=1, how='all')

        # Change column names using range
        qas.columns = range(len(qas.columns))
        normal_samples.columns = range(len(normal_samples.columns))
        qas2.columns = range(len(qas2.columns))
        samples2.columns = range(len(samples2.columns))

        # Concatenate samples and qas
        samples_and_qa = pd.concat([qas, normal_samples], ignore_index=True, axis=0)
        samples_and_qa2 = pd.concat([qas2, samples2], ignore_index=True, axis=0)
        normal_samples = pd.concat([samples_and_qa, samples_and_qa2], ignore_index=True, axis=1).dropna(axis=1,
                                                                                                        how='all')
        # ? Only god and I knew what this does... And I forgot.
        normal_samples = normal_samples.transpose()
        normal_samples = normal_samples.melt(id_vars=[0, 1, 2, 3, 4, 5], ignore_index=False).sort_index()
        normal_samples['index'] = normal_samples.index

        normal_samples.set_index(['index', 'variable'], inplace=True)
        normal_samples = normal_samples.sort_index()
        normal_samples["ri_log_id"] = log_df['ri_log_id'].iloc[0]

        normal_samples.columns = ['qa', 'characteristic', 'nominal_value', 'superior_limit', 'inferior_limit',
                                  'equipment', 'sample_value', 'ri_log_id']

        normal_samples.reset_index(inplace=True)
        normal_samples.rename(columns={'index': 'sample_batch', 'variable': 'sample_number'}, inplace=True)

        return log_df, normal_samples

    # If there's only one...
    elif len(qa_indexes) == 1:
        # Slice the DataFrame using loc up to the QA#, but not including it.
        row = df.loc[qa_indexes[0] - 1:, :]

        # Separate samples from qas
        qas = row.iloc[:6, 1::2]
        normal_samples = row.iloc[6:, :].dropna(axis=1, how='all')

        # Change column names using range
        qas.columns = range(len(qas.columns))
        normal_samples.columns = range(len(normal_samples.columns))

        # Concatenate samples and qas
        samples_and_qa = pd.concat([qas, normal_samples], ignore_index=True, axis=0)
        normal_samples = pd.concat([samples_and_qa], ignore_index=True, axis=1).dropna(axis=1, how='all')

        normal_samples = normal_samples.transpose()
        normal_samples = normal_samples.melt(id_vars=[0, 1, 2, 3, 4, 5], ignore_index=False).sort_index()
        normal_samples['index'] = normal_samples.index

        normal_samples.set_index(['index', 'variable'], inplace=True)
        normal_samples = normal_samples.sort_index()

        normal_samples["ri_log_id"] = log_df['ri_log_id'].iloc[0]

        normal_samples.columns = ['qa', 'characteristic', 'nominal_value', 'superior_limit', 'inferior_limit',
                                  'equipment', 'sample_value', 'ri_log_id']

        normal_samples.reset_index(inplace=True)
        normal_samples.rename(columns={'index': 'sample_batch', 'variable': 'sample_number'}, inplace=True)

        return log_df, normal_samples

    else:
        return None, None


def read_report(report_path, kind):
    """
    Reads an Excel report and returns the log and samples for the report type.

    Args:
        kind (str): Whether is normal or special
        report_path (str): Full path to the Excel report.

    Returns:
        dict: Dictionary with report type, log, and samples.
    """

    try:
        report_wb = openpyxl.load_workbook(report_path, data_only=True, read_only=True)
        report_ws = report_wb.active

        # * Finds if its "Concentricidad" or "Ovalidad"
        if report_ws["N6"].value == "CONC":
            log, samples = read_concentricity(report_ws, kind)
            return {'Type': "CONC", "log": log, "Samples": samples}

        elif report_ws["D6"].value == "OVAL":
            log, samples = read_ovality(report_ws, kind)
            return {'Type': "OVAL", "log": log, "Samples": samples}

        else:
            log, samples = read_normal(report_ws, kind)
            return {'Type': "OTHER", 'log': log, 'Samples': samples}

    except PermissionError:
        print("\rPermission denied: {}".format(report_path))
        return
    except FileNotFoundError:
        print("\rFile not found: {}".format(report_path))
        return
    except zipfile.BadZipFile:
        print("\rNot a valid Excel file: {}".format(report_path))
        return
    except InvalidFileException:
        print("\rNot a valid Excel file: {}".format(report_path))


def change_start_id(df: pd.DataFrame, start_id: int):
    """
    This function changes the index of a Dataframe to continue from any given number. i.e:
    start_id = 9; new_id = 10 ...

    Args:
        df (pd.DataFrame):  Target dataframe
        start_id (int): The index that the dataframe will begin with.
    Returns:
        pd.DataFrame
    """

    df.index = range(start_id + 1, start_id + len(df) + 1)

    return df


def clean_numbers(column_to_modify):
    clean_column = column_to_modify
    clean_column = clean_column.astype(str)
    clean_column = clean_column.str.extract(r'(\d+)')
    clean_column = clean_column.fillna(0)
    clean_column = clean_column.astype(int)
    return clean_column


def scrape_ri(kind: str, report_directory: str, save_to: str, upload=True, testing='', samples=100):
    """
    Scrape data from RI reports and save it to CSV and SQL Server.

    Args:
        kind (str): The kind of RI report.
        report_directory (str): The directory containing all RI reports.
        save_to (str): The directory to save the output files.
        upload (bool): Whether to upload the data to SQL Server. Default is True.
        testing (str): Whether to run the function in test mode. Default is False.
        samples (int): The number of samples to use in testing mode. Default is 100.
    """

    reports_list = []
    ri_logs = pd.DataFrame()
    ri_tests = pd.DataFrame()
    count = 0
    logs_table_name = fr'ri_logs'
    samples_table_name = fr'ri_tests'

    # Use the reports_list to analyze whether the files are already on the db or not.
    # If they are, then drop those files, so they don't get scanned, saving valuable time.
    with sql_engine.connect() as conn:
        result = conn.execute(text(f"SELECT ri_log_id FROM {logs_table_name}"))
        db_log_ids = [str(value[0]) for value in result.all()]

    for path, subdirs, files in os.walk(report_directory):
        for file in files:
            if file.endswith('xlsm'):
                file_name = file.split('.')[0]
                if file_name not in db_log_ids:  # ? if ???? and 'Reporte' in file_name
                    file_path = os.path.join(path, file)
                    file_size = os.path.getsize(file_path)
                    if file_size < 1_000_000:  # * Less than a Mb
                        reports_list.append(file_path)

    total_files = len(reports_list)
    print('Files to be uploaded: ', total_files)

    with concurrent.futures.ProcessPoolExecutor() as executor:
        if testing == 'random':
            results = executor.map(read_report, random.choices(reports_list, k=samples), repeat(kind))
        if testing == 'first_n':
            results = executor.map(read_report, reports_list[:samples], repeat(kind))
        else:
            results = executor.map(read_report, reports_list, repeat(kind))

        for result in results:
            if result is not None:
                count += 1
                # ? wtf is this?
                # {result["log"]["ri_log_id"].iloc[0]}
                print(f'\rFiles scanned: {count}', end='')
                ri_logs = pd.concat([ri_logs, result['log']], ignore_index=True)
                ri_tests = pd.concat([ri_tests, result['Samples']], ignore_index=True)

    ri_tests.rename_axis(index={None: "id"}, inplace=True)

    # * Cleaning the data before exporting it.
    if not ri_logs.empty:
        ri_logs.rename_axis(index={None: "id"}, inplace=True)
        ri_logs['quantity'] = clean_numbers(ri_logs['quantity'])
        ri_logs['shift'] = clean_numbers(ri_logs['shift'])
        ri_logs['ri_log_id'] = clean_numbers(ri_logs['ri_log_id'])
        ri_logs.dropna(subset=['ri_log_id'], inplace=True)
        ri_logs.drop_duplicates(subset=['ri_log_id'], inplace=True)
        ri_logs['technician'] = ri_logs['technician'].str.title()
        ri_logs['line'] = ri_logs['line'].astype(str)
        ri_logs['log_date'] = pd.to_datetime(ri_logs['log_date'], errors='coerce')
        ri_logs.dropna(subset='log_date', inplace=True)
        ri_logs.replace('N/A', None, inplace=True)
        ri_logs['kind'] = kind

    if not ri_tests.empty:
        ri_tests['sample_number'] = pd.to_numeric(ri_tests['sample_number'], errors='coerce')
        ri_tests['sample_batch'] = pd.to_numeric(ri_tests['sample_batch'], errors='coerce')
        ri_tests['ri_log_id'] = clean_numbers(ri_tests['ri_log_id'])
        ri_tests.dropna(subset='ri_log_id', inplace=True)

        for column_name in ['nominal_value', 'superior_limit', 'inferior_limit']:
            ri_tests[column_name] = ri_tests[column_name].astype(str).str.extract(r'(\d+\.\d+|\d+)').astype(float).fillna(0.0)

        # TODO. Check if this works ...                                                          ↓↓↓↓↓↓
        ri_tests['sample_value'].astype(str).str.extract(r'(\d+\.\d+|\d+)').astype(float).fillna(np.nan)

    # Todo. drop Supplier column & Get Line and Shift from matrix file.

    # * Saving the data into .csv files.
    ri_logs.to_csv(save_to + fr'\\{logs_table_name}.csv')
    ri_tests.to_csv(save_to + fr'\\{samples_table_name}.csv')
    print('\nFiles saved.')

    if upload and not ri_logs.empty and not ri_tests.empty:
        # * Drop ri_logs.ri_log_id that are already in the db.
        log_duplicates = [value for value in ri_logs['ri_log_id'] if str(value) in db_log_ids]
        ri_logs = ri_logs[~ri_logs['ri_log_id'].isin(log_duplicates)]

        ri_logs.to_sql('ri_logs', sql_engine, if_exists='append', index=False, dtype=log_dtypes)
        ri_tests = ri_tests[ri_tests['ri_log_id'].isin(ri_logs['ri_log_id'].tolist())]
        ri_tests.to_sql('ri_tests', sql_engine, if_exists='append', index=False, dtype=tests_dtypes)
        print('Logs uploaded: ', len(ri_logs))
        print('Samples uploaded: ', len(ri_tests))
    else:
        print("No new data found, didn't upload any files.")

if __name__ == '__main__':
    scrape_ri(
        kind='special',
        report_directory=special_path_2023,
        save_to=output_path,
        upload=True
    )
